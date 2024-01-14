import openpyxl
import pandas as pd
import D1g1tObject


# Ready to test
def load_files(latest_export_path, master_path):
    """
    Load the two Excel files and return them in a tuple

    Args:
        latest_export_path: Path to the latest d1g1t export.
        master_path: Path to the master.

    Returns:
        tuple: A tuple containing excel files.
    """
    latest_export = openpyxl.load_workbook(latest_export_path)
    master = openpyxl.load_workbook(master_path)
    return latest_export, master


# Ready to test
def load_exceptions(exception_path):
    """
    Load the exceptions file and return it.

    Args:
        exception_path: Path to the exceptions file.

    Returns:
        pandas.DataFrame: The exceptions DataFrame.
    """
    exceptions = pd.read_excel("../Macro exceptions.xlsx")  # .dropna()
    exceptions = exceptions.fillna("")

    return exceptions


# Ready to test
def generate_d1g1t_objects(workbook):
    """
    Generate D1g1t objects from an Excel file.

    Args:
        file_path (str): Path to the Excel file.

    Returns:
        dict: A dictionary of D1g1t objects.
    """
    sheet = workbook.active  # sheet = active sheet in workbook
    d1g1t_objects = {}  # d1g1t_objects = dictionary of D1g1t Objects

    # Looping through each row to create a D1g1t Object
    for i in range(2, sheet.max_row + 1):
        ident = sheet["A" + str(i)].value  # ident = Security ID or Account ID in master
        attributes = {}  # attributes = dictionary of attributes for each D1g1t Object
        for j in range(0, sheet.max_column):
            k = sheet[chr(j + 65) + str(1)].value  # k = column header
            v = sheet[chr(j + 65) + str(i)].value  # v = attribute value
            if v == None or v == "Undefined":
                v = ""
            attributes[k] = v

        # Create D1g1tObject
        obj = D1g1tObject(ident, attributes)

        # Add D1g1tObject to Dictionary
        d1g1t_objects[ident] = obj

    # Return dictionary of D1g1t Objects
    return d1g1t_objects


# TODO: Write function
def compare_workbooks(latest_export_objects, master_objects):
    """
    Compare two workbooks and return a list of changes.

    Args:
        latest_export_objects (dict): A dictionary of D1g1t objects from the latest export.
        master_objects (dict): A dictionary of D1g1t objects from the master.

    Returns:
        list: A list of changes.
    """
    changes = []
    for ident in latest_export_objects:
        latest_export_obj = latest_export_objects[ident]
        master_obj = master_objects[ident]
        if latest_export_obj != master_obj:
            changes.append((latest_export_obj, master_obj))
    return changes


# TODO: Write function
def add_sheet_to_master():
    """
    Add a sheet to the master.

    Args:
        None

    Returns:
        None
    """
    pass


# TODO: Write function
def main():
    """
    Main function to execute the program.

    Asks the user for file paths, loads the files, and creates a change file.

    Returns:
        None
    """
    file1_path = input("Enter path to original master:\n")
    file2_path = input("Enter path to latest d1g1t export:\n")

    df1, df2 = load_files(file1_path, file2_path)
