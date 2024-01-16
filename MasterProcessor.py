import openpyxl
import pandas as pd
from D1g1tObject import D1g1tObject
from datetime import date
from yfinance import Ticker

# LOAD EXCEPTIONS


class MasterProcessor:
    """
    A class to represent a D1g1t object.

    Attributes:
        ident (str): The identifier of the D1g1t object.
        fields_dict (dict): A dictionary of fields and their values.
    """

    def __init__(self, master_filepath, latest_d1g1t_export_filepath):
        """
        Constructor for the D1g1tObject class.

        Args:
            ident (str): The identifier of the D1g1t object.
            fields_dict (dict): A dictionary of fields and their values.
        """
        self.master_filepath = master_filepath
        self.latest_d1g1t_export_filepath = latest_d1g1t_export_filepath

        self.exceptions_df = pd.read_excel(r"Macro exceptions.xlsx")
        self.exceptions_df = self.exceptions_df.fillna("")
        self.exceptions = self.exceptions_df.set_index("Find").T.to_dict("list")

    def run_master_processor(self):
        master, latest_d1g1t_export = self.load_files(
            self.master_filepath, self.latest_d1g1t_export_filepath
        )  # WORKING

        headers = self.generate_headers(self.latest_d1g1t_export_filepath)

        master_objects, latest_d1g1t_export_objects = self.generate_object_dicts_for_comparison(
            master, latest_d1g1t_export
        )

        new_objects, deleted_objects, surviving_objects = self.group_objects(
            latest_d1g1t_export_objects, master_objects
        )

        self.prepare_new_master(headers, self.master_filepath)

        # TODO: DEBUG
        self.compare_object_dictionaries(
            self.master_filepath, headers, surviving_objects, master_objects, latest_d1g1t_export_objects
        )
        self.add_new_objects_to_master(
            self.master_filepath, headers, new_objects, latest_d1g1t_export_objects
        )

    # WORKING
    def load_files(self, master_path, latest_export_path):
        """
        Load the two Excel files and return them in a tuple

        Args:
            latest_export_path: Path to the latest d1g1t export.
            master_path: Path to the master.

        Returns:
            tuple: A tuple containing excel files.
        """
        master = openpyxl.load_workbook(master_path)
        latest_export = openpyxl.load_workbook(latest_export_path)
        return master, latest_export

    # Ready to test
    def load_exceptions(self, exception_path):
        """
        Load the exceptions file and return it.

        Args:
            exception_path: Path to the exceptions file.

        Returns:
            pandas.DataFrame: The exceptions DataFrame.
        """
        exceptions = pd.read_excel("../Macro exceptions.xlsx")  # .dropna()
        exceptions = exceptions.fillna("")
        exceptions = self.exceptions_df.set_index("Find").T.to_dict("list")

        return exceptions

    # Ready to test
    def generate_d1g1t_objects(self, workbook):
        """
        Generate D1g1t objects from an Excel file.

        Args:
            file_path (str): Path to the Excel file.

        Returns:
            dict: A dictionary of D1g1t objects.
        """
        # workbook = openpyxl.load_workbook(workbook)
        sheet = workbook.active  # sheet = active sheet in workbook
        d1g1t_objects = {}  # d1g1t_objects = dictionary of D1g1t Objects

        # Looping through each row to create a D1g1t Object
        for i in range(2, sheet.max_row + 1):
            ident = sheet["A" + str(i)].value  # ident = Security ID or Account ID in master
            attributes = {}  # attributes = dictionary of attributes for each D1g1t Object
            for j in range(0, sheet.max_column):
                k = sheet[chr(j + 65) + str(1)].value  # k = column header
                v = sheet[chr(j + 65) + str(i)].value  # v = attribute value
                if v is None or v == "Undefined":
                    v = ""
                attributes[k] = v

            # Create D1g1tObject
            obj = D1g1tObject(ident, attributes)

            # Add D1g1tObject to Dictionary
            d1g1t_objects[ident] = obj

        # Return dictionary of D1g1t Objects
        return d1g1t_objects

    # Ready to test
    def generate_object_dicts_for_comparison(self, master, latest_d1g1t_export):
        """
        Generate dictionaries of D1g1t objects from the master and latest export.

        Args:
            master (openpyxl.workbook): The master workbook.
            latest_d1g1t_export (openpyxl.workbook): The latest export from D1g1t.

        Returns:
            tuple: A tuple containing dictionaries of D1g1t objects.
        """
        master_objects = self.generate_d1g1t_objects(master)
        latest_d1g1t_export_objects = self.generate_d1g1t_objects(latest_d1g1t_export)
        return master_objects, latest_d1g1t_export_objects

    # Ready to test
    def group_objects(self, latest_d1g1t_export_objects, latest_master_objects):
        # Objects that are new since the last export. To be highlighted in LIGHT GREEN
        new_objects = set(latest_d1g1t_export_objects.keys()).difference(set(latest_master_objects.keys()))

        # Objects that have been deleted since the last export. To be listed in RunSummary
        deleted_objects = set(latest_master_objects.keys()).difference(
            set(latest_d1g1t_export_objects.keys())
        )

        # Objects that are in both the latest export and the master
        surviving_objects = set(latest_d1g1t_export_objects.keys()) - new_objects

        return new_objects, deleted_objects, surviving_objects

    def generate_headers(self, latest_d1g1t_export_filepath):
        """
        Generate headers from the latest d1g1t export file.

        Args:
            latest_d1g1t_export_filepath (str): The file path of the latest d1g1t export file.

        Returns:
            list: A list of headers extracted from the export file.
        """
        headers = []
        latest_d1g1t_export_workbook = openpyxl.load_workbook(latest_d1g1t_export_filepath)
        latest_d1g1t_export_sheet = latest_d1g1t_export_workbook.active
        for i in range(0, latest_d1g1t_export_sheet.max_column):
            headers.append(latest_d1g1t_export_sheet[chr(65 + i) + "1"].value)
        return headers

    def mark_as_updated(self, cell):
        """
        Mark a cell as updated by changing the background color to YELLOW.

        Args:
            cell (openpyxl.cell): The cell to be marked as updated.

        Returns:
            None
        """
        cell.fill = openpyxl.styles.PatternFill(patternType="solid", fgColor="FFFF00")

    def mark_as_new(self, cell):
        """
        Mark a cell as new by changing the background color to LIGHT GREEN.

        Args:
            cell (openpyxl.cell): The cell to be marked as new.

        Returns:
            None
        """
        cell.fill = openpyxl.styles.PatternFill(patternType="solid", fgColor="7FFFD4")

    def mark_as_generated(self, cell):
        """
        Mark a cell as generated by changing the background color to LIGHT BLUE.

        Args:
            cell (openpyxl.cell): The cell to be marked as generated.

        Returns:
            None
        """
        cell.fill = openpyxl.styles.PatternFill(patternType="solid", fgColor="ADD8E6")

    def prepare_new_master(self, headers, latest_master_path):
        """
        Creates a new sheet in the master workbook and adds headers to it.

        Args:
            headers (list): List of headers to be added to the new sheet.
            latest_master (str): Path of the latest master workbook.

        Returns:
            None
        """
        # Create a new sheet in the master workbook
        master = openpyxl.load_workbook(latest_master_path)
        master_sheet = master.create_sheet(date.today().strftime("%b %d"), 0)

        # Add headers to the new sheet (consistent headers with latest export)
        for i in range(0, len(headers)):
            master_sheet[chr(65 + i) + "1"] = headers[i]

        # Save the new master workbook
        master.save(latest_master_path)

    def clean_name(self, orig_name):
        if orig_name is not None:
            name = str(orig_name).split()  # Generates a list of tokens

            for i in range(len(name)):
                name[i] = name[i].replace("*", "")
                name[i] = name[i].title()  # Makes the first letter of every word capitalized
                name[i] = name[i].replace("'S", "'s")
                name[i] = name[i].replace("Wts-", "WTS ")
                name[i] = name[i].replace(".Com", ".com")
                if name[i] in self.exceptions.keys():  # Checks if token triggers an exception
                    name[i] = str(self.exceptions[name[i]][0])  # If yes, replaces with new exception

            name = " ".join(name).strip()
            # name = " ".join(str(name)).strip()              #Joins tokens and returns clean name
            return name

        else:
            return ""

    def get_security_asset_subclass(self, ticker):
        # Get ticker information, market cap and price-to-sales ratio
        info = Ticker(ticker).info
        market_cap = info.get("marketCap", 0)
        price_to_sales = info.get("trailingPE", 0)

        size = "Large Cap" if market_cap >= 10e9 else "Mid Cap" if 2e9 <= market_cap < 10e9 else "Small Cap"
        category = "Growth" if price_to_sales > 15 else "Blend" if price_to_sales == 15 else "Value"

        return size + " " + category

    def compare_object_dictionaries(
        self, master_filepath, headers, surviving_objects, latest_master_objects, latest_d1g1t_export_objects
    ):
        master = openpyxl.load_workbook(master_filepath)
        master_sheet = master.active  # Confirm this is the new sheet

        # Loop through the surviving objects
        surviving_objects = list(surviving_objects)
        i = 2
        for obj in surviving_objects:
            for j in range(0, len(headers)):
                cell = master_sheet[chr(65 + j) + str(i)]
                old = ""
                if headers[j] in set(latest_master_objects[obj].fields_dict.keys()):
                    old = latest_master_objects[obj].fields_dict[headers[j]]
                if headers[j] == "Security Name":
                    new = self.clean_name(latest_d1g1t_export_objects[obj].fields_dict[headers[j]])
                    new = " ".join(new.split())
                    # Above turns names such as "Put/Xsp                 @  378 Exp 07/21/2023" into "Put/Xsp @ 378 Exp 07/21/2023"
                else:
                    new = latest_d1g1t_export_objects[obj].fields_dict[headers[j]]
                cell.value = new
                if old != new:
                    self.mark_as_updated(cell)
                else:
                    no_fill = openpyxl.styles.PatternFill(fill_type=None)
                    master_sheet[chr(65 + j) + str(i)].fill = no_fill
            i = i + 1
        i = i - 1
        master.save(master_filepath)

    def add_new_objects_to_master(self, master_filepath, headers, new_objects, latest_d1g1t_export_objects):
        """
        Adds new objects to the master file.

        Args:
            master_filepath (str): The file path of the master file.
            headers (list): The list of headers for the new objects.
            new_objects (list): The list of new objects to be added.
            latest_d1g1t_export_objects (dict): The dictionary containing the latest D1G1T export objects.

        Returns:
            None
        """
        master = openpyxl.load_workbook(master_filepath)
        master_sheet = master.active
        new_objects = list(new_objects)
        ticker = ""
        i = master_sheet.max_row + 1
        for obj in new_objects:
            for j in range(0, len(headers)):
                cell = master_sheet[chr(65 + j) + str(i)]
                if headers[j] == "Security Name":
                    new = self.clean_name(latest_d1g1t_export_objects[obj].fields_dict[headers[j]])
                    # Above turns names such as "Put/Xsp                 @  378 Exp 07/21/2023" into "Put/Xsp @ 378 Exp 07/21/2023"
                else:
                    new = latest_d1g1t_export_objects[obj].fields_dict[headers[j]]
                if new != "Undefined":
                    cell.value = new  # add empty string below
                    self.mark_as_new(cell)
                # master_sheet[chr(65 + j) + str(i)].fill = fill_pattern_if_new
                if headers[j] == "Ticker Short":
                    ticker = master_sheet[chr(65 + j) + str(i)].value
                if headers[j] == "Sector":
                    if ticker is not None:
                        try:
                            ticker_info = Ticker(ticker).info
                            sector = ticker_info["sector"]
                            if sector == "Basic Materials":
                                sector = "Materials"
                            print(sector)
                            master_sheet[chr(65 + j) + str(i)].value = sector
                            self.mark_as_generated(cell)
                        except Exception as e:
                            print("Ticker unavailable", e)
                if headers[j] == "Security Asset Sub-class":
                    if ticker is not None:
                        try:
                            sasc = self.get_security_asset_subclass(
                                ticker
                            )  # Get ticker information, market cap and price-to-sales ratio
                            self.mark_as_generated(cell)
                            cell.value = sasc
                        except Exception as e:
                            cell.value = latest_d1g1t_export_objects[obj].fields_dict[headers[j]]
                            print(new_objects[obj].fields_dict[headers[j]])
                            print("SASC unavailable", e)
            i = i + 1
        i = i - 1
        master.save(master_filepath)
