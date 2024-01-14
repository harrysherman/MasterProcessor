from json import load
import pandas as pd
import os
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

# from D1g1tObject import D1g1tObject
from datetime import date
import yfinance

today = date.today()
today_as_string = today.strftime("%b_%d_%y")

exceptions_df = pd.read_excel("../Macro exceptions.xlsx")  # .dropna()
exceptions_df = exceptions_df.fillna("")


# create a dictionary mapping "find" to "change to"
exceptions = exceptions_df.set_index("Find").T.to_dict("list")


headers = []


class D1g1tObject:
    def __init__(self, id, fields_dict):
        self.id = id
        self.fields_dict = fields_dict


"""Essentially our "main()" function."""


def generate_object_lists():
    # Open Latest Master and build list of D1g1tObjects
    filename = ""
    cur_path = os.getcwd()
    path_to_master = os.path.dirname(os.getcwd())
    os.chdir(path_to_master)
    files = os.listdir()
    for f in files:
        if "Securities Master" in f or "securities master" in f:
            filename = f
    # latest_master_workbook = openpyxl.load_workbook(path_to_master+"/"+filename)
    latest_master_workbook = openpyxl.load_workbook(os.path.join(path_to_master, filename))

    """Add Sheet to existing master!"""
    latest_master_sheet = latest_master_workbook.active
    latest_master_objects = generate_d1g1t_objects(latest_master_sheet)

    # Go back to original location
    os.chdir(cur_path)

    # Open Latest D1G1T Export and build list of D1g1tObjects
    filename = ""
    files = os.listdir("../Latest D1G1T Export/")
    for f in files:
        if ".DS_Store" not in f:
            filename = f
    latest_d1g1t_export_workbook = openpyxl.load_workbook("../Latest D1G1T Export/" + filename)
    latest_d1g1t_export_sheet = latest_d1g1t_export_workbook.active
    latest_d1g1t_export_objects = generate_d1g1t_objects(latest_d1g1t_export_sheet)
    for i in range(0, latest_d1g1t_export_sheet.max_column):
        headers.append(latest_d1g1t_export_sheet[chr(65 + i) + "1"].value)
    # TODO: Delete this sheet

    return latest_master_objects, latest_d1g1t_export_objects


"""Begin populating and looking for differences"""


def populate_new_master(latest_master_objects, latest_d1g1t_export_objects):
    path_to_master = os.path.dirname(os.getcwd())
    os.chdir(path_to_master)

    files = os.listdir()
    for f in files:
        if "Securities Master" in f or "securities master" in f:
            filename = f
    master = openpyxl.load_workbook(os.path.join(path_to_master, filename))
    master_sheet = master.create_sheet(date.today().strftime("%b %d"), 0)

    # Isolate NEW objects
    new_objects = set(latest_d1g1t_export_objects.keys()).difference(set(latest_master_objects.keys()))
    deleted_objects = set(latest_master_objects.keys()).difference(set(latest_d1g1t_export_objects.keys()))
    surviving_objects = set(latest_d1g1t_export_objects.keys()) - new_objects

    # Add headers to master
    for i in range(0, len(headers)):
        master_sheet[chr(65 + i) + "1"] = headers[i]

    # Initialize fill pattern
    fill_pattern_if_changed = PatternFill(patternType="solid", fgColor="ffff00")

    # populate and compare
    objects = list(surviving_objects)
    i = 2
    for obj in objects:
        for j in range(len(headers)):
            old = ""
            if headers[j] in set(latest_master_objects[obj].fields_dict.keys()):
                old = latest_master_objects[obj].fields_dict[headers[j]]
            if headers[j] == "Security Name":
                new = clean_name(latest_d1g1t_export_objects[obj].fields_dict[headers[j]])
                new = " ".join(new.split())
                # Above turns names such as "Put/Xsp                 @  378 Exp 07/21/2023" into "Put/Xsp @ 378 Exp 07/21/2023"
            else:
                new = latest_d1g1t_export_objects[obj].fields_dict[headers[j]]
            cell = master_sheet[chr(65 + j) + str(i)]
            cell.value = new
            if old != new:
                master_sheet[chr(65 + j) + str(i)].fill = fill_pattern_if_changed
            else:
                no_fill = openpyxl.styles.PatternFill(fill_type=None)
                master_sheet[chr(65 + j) + str(i)].fill = no_fill

        i = i + 1
    i = i - 1

    # Add new objects #BEWARE OF SCOPE OF I WHILE TESTING
    fill_pattern_if_new = PatternFill(patternType="solid", fgColor="7FFFD4")
    fill_pattern_if_generated = PatternFill(patternType="solid", fgColor="FFA500")
    new_objects = list(new_objects)
    ticker = ""
    for obj in new_objects:
        for j in range(len(headers)):
            """CLEAN NAMES OF NEW SECS"""
            if headers[j] == "Security Name":
                new = clean_name(latest_d1g1t_export_objects[obj].fields_dict[headers[j]])
                # Above turns names such as "Put/Xsp                 @  378 Exp 07/21/2023" into "Put/Xsp @ 378 Exp 07/21/2023"
            else:
                new = latest_d1g1t_export_objects[obj].fields_dict[headers[j]]
            cell = master_sheet[chr(65 + j) + str(i)]
            if new != "Undefined":
                cell.value = new  # add empty string below
            master_sheet[chr(65 + j) + str(i)].fill = fill_pattern_if_new
            # ticker = ''
            if headers[j] == "Ticker Short":
                ticker = master_sheet[chr(65 + j) + str(i)].value
            if headers[j] == "Sector":
                if ticker is not None:
                    try:
                        ticker_info = yfinance.Ticker(ticker).info
                        sector = ticker_info["sector"]
                        print(sector)
                        if sector == "Basic Materials":
                            sector = "Materials"
                        master_sheet[chr(65 + j) + str(i)].value = sector
                        master_sheet[chr(65 + j) + str(i)].fill = fill_pattern_if_generated
                    except:
                        print("Ticker unavailable")
            if headers[j] == "Security Asset Sub-class":
                if ticker is not None:
                    try:
                        # Get ticker information, market cap and price-to-sales ratio
                        info = yfinance.Ticker(ticker).info
                        market_cap = info.get("marketCap", 0)
                        price_to_sales = info.get("trailingPE", 0)

                        # Determine size (Large Cap, Mid Cap, Small Cap) and category (Growth, Value, Blend)
                        size = (
                            "Large Cap"
                            if market_cap >= 10e9
                            else "Mid Cap"
                            if 2e9 <= market_cap < 10e9
                            else "Small Cap"
                        )
                        category = (
                            "Growth" if price_to_sales > 15 else "Blend" if price_to_sales == 15 else "Value"
                        )

                        # Return Security Asset Sub Class
                        master_sheet[chr(65 + j) + str(i)].fill = fill_pattern_if_generated
                        cell.value = size + " " + category
                    except:
                        cell.value = latest_d1g1t_export_objects[obj].fields_dict[headers[j]]
                        print(latest_d1g1t_export_objects[obj].fields_dict[headers[j]])

        i = i + 1
    i = i - 1

    # Add Deleted Objects
    fill_pattern_if_deleted = PatternFill(patternType="solid", fgColor="FF3333")
    new_objects = list(new_objects)
    for obj in deleted_objects:
        for j in range(len(headers)):
            if headers[j] in list(latest_master_objects[obj].fields_dict.keys()):
                new = latest_master_objects[obj].fields_dict[headers[j]]
                cell = master_sheet[chr(65 + j) + str(i)]
                cell.value = new
            master_sheet[chr(65 + j) + str(i)].fill = fill_pattern_if_deleted
        i = i + 1

    # Create Table
    row = str(master_sheet.max_row)
    col = chr(65 + len(headers) - 1)
    table = Table(displayName="Table" + today_as_string, ref="A1:" + col + row)

    master_sheet.add_table(table)

    master.save("Securities Master.xlsx")

    return master


"""Helper method to generate dictionary of d1g1t objects"""


def generate_d1g1t_objects(sheet):
    d1g1t_objects = {}

    for i in range(2, sheet.max_row + 1):
        ident = sheet["A" + str(i)].value
        atr_dict = {}
        for j in range(0, sheet.max_column):
            k = sheet[chr(j + 65) + str(1)].value  # k = column header
            v = sheet[chr(j + 65) + str(i)].value  # v = field value
            if v == None or v == "Undefined":
                v = ""
            atr_dict[k] = v

        # Create D1g1tObject
        obj = D1g1tObject(ident, atr_dict)

        # Add D1g1tObject to the list
        d1g1t_objects[ident] = obj

    # Return list of D1g1t Objects
    return d1g1t_objects


def test():
    return ""


# def load_exceptions():
#    exceptions_df = pd.read_excel(os.path.join("Master Processing/","Macro exceptions.xlsx")).dropna()

# create a dictionary mapping "find" to "change ot"
#   exceptions = exceptions_df.set_index("Find").T.to_dict('list')

#  return exceptions

# exceptions = load_exceptions()


def clean_name(orig_name):
    if orig_name is not None:
        name = str(orig_name).split()  # Generates a list of tokens

        for i in range(len(name)):
            name[i] = name[i].replace("*", "")
            name[i] = name[i].title()  # Makes the first letter of every word capitalized
            name[i] = name[i].replace("'S", "'s")
            name[i] = name[i].replace("Wts-", "WTS ")
            name[i] = name[i].replace(".Com", ".com")
            if name[i] in exceptions.keys():  # Checks if token triggers an exception
                name[i] = str(exceptions[name[i]][0])  # If yes, replaces with new exception

        name = " ".join(name).strip()
        # name = " ".join(str(name)).strip()              #Joins tokens and returns clean name
        return name

    else:
        return ""


if __name__ == "__main__":
    latest_master_objects, latest_d1g1t_export_objects = generate_object_lists()
    populate_new_master(latest_master_objects, latest_d1g1t_export_objects)
